import openpyxl


# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)
